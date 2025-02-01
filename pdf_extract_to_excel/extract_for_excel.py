import boto3  # Import the boto3 library for AWS services
import botocore  # Import the botocore library for AWS low-level service operations
import json  # Import the json library for working with JSON data
import streamlit as st  # Import the Streamlit library for building web apps
import pdfplumber  # Import the pdfplumber library for PDF extraction
from dotenv import load_dotenv  # Import the load_dotenv function from the dotenv library
import os  # Import the os module for interacting with the operating system  # Import the pdf_processing function from the extract_pdf_to_json module
from excel_prompt_page_duree import excel_extraction_page_duree
from excel_prompt_page_charge import excel_extraction_page_charge
#from excel_prompt_page_divers import excel_extraction_page_divers

# Loading environment variables from the .env file
load_dotenv()

# Setting up the default AWS session using the profile name from the environment variable
boto3.setup_default_session(profile_name=os.getenv("hackathon"))

# Setting up the Bedrock client with a custom timeout configuration
config = botocore.config.Config(connect_timeout=300, read_timeout=300)
bedrock = boto3.client('bedrock-runtime', 'us-west-2', config=config)

test_pdf_path = "/Users/gk/Downloads/HarmonieBail.pdf"

def parse_xml(xml, tag):
    """
    Helper function to extract the value between XML tags from a given string.
    """
    start_tag = f"<{tag}>"
    end_tag = f"</{tag}>"
    
    start_index = xml.find(start_tag)
    if start_index == -1:
        return ""

    end_index = xml.find(end_tag)
    if end_index == -1:
        return ""

    value = xml[start_index+len(start_tag):end_index]
    return value

def extract_from_pdf(pdf_file, to_do):
    text = ""
    print("processing pdf...")
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text += page.extract_text()
    print("pdf processed !")

    result = {}

    print("processing tags one by one")
    for thing in to_do:
        print("processing", thing[0]["tag"])
        input = [{
        "type": "text",
        "text": text
        }]
        for call in thing:
            prompt = {
                "temperature": 0.0,
                "messages": [  
                    {
                        "role": "system",
                        "content": call["sys_prompt"]
                    },
                    {
                        "role": "user",
                        "content": f"<legal_call_content> {input} </legal_call_content>"
                    }
                ]
            }
            prompt = json.dumps(prompt)
            response = bedrock.invoke_model(body=prompt, modelId="mistral.mistral-large-2407-v1:0", accept="application/json", contentType="application/json")
            response_body = json.loads(response.get('body').read())
            llmOutput = response_body['choices'][0]['message']['content']
            print("llmOutput", llmOutput)
            output = parse_xml(llmOutput, "output")
            input = [{"type": "text", "text": output}]
        result[thing[0]["tag"]] = input[0]["text"]
    return result

#extract the data from the pdf
result_dict1 = extract_from_pdf(test_pdf_path, excel_extraction_page_charge)
result_dict2 = {'ACTIF': result_dict1['ACTIF']} 
result_dict2.update(extract_from_pdf(test_pdf_path, excel_extraction_page_duree))
#result_dict3 = extract_from_pdf(test_pdf_path, excel_extraction_page_divers)

from openpyxl import load_workbook
wb = load_workbook(filename ="Tableau_audit_immo_modele.xlsx")
output_file = 'Tableau_audit_immo.xlsx'
#get the excel sheets
ws1 = wb["Durée - Loyers"]
ws2 = wb["Refacturation des charges"]
ws3 = wb["Divers"]

def populate_excel(ws, data):
    i = 3
    while ws.cell(row=i, column=3).value is not None:
        i += 1
    for column_index, tag  in enumerate(data):
        ws.cell(row=i, column=column_index+2).value = data[tag]
    return "done"
#populate the excel sheet with the extracted data
populate_excel(ws1, result_dict1)
populate_excel(ws2, result_dict2)
#populate_excel(ws3, result_dict3)

wb.save(output_file)
if __name__ == "__main__":
    with open(test_pdf_path, 'rb') as file:
        print(extract_from_pdf(file, excel_extraction_page_duree))