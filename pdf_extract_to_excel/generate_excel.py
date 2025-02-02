from pdf_extract_to_excel.extract_for_excel import extract_from_pdf, extract_from_pdf_fast
from openpyxl import load_workbook
from pdf_extract_to_excel.excel_prompt_page_duree import excel_extraction_page_duree
from pdf_extract_to_excel.excel_promp_page_divers import excel_extraction_page_divers
from pdf_extract_to_excel.excel_prompt_page_charge import excel_extraction_page_charge
from pdf_extract_to_excel.excel_tag_column_correspondance import correspondance_duree, correspondance_charges, correspondance_divers

test_pdf_path = "C:\\Users\\mateo\\Documents\\ecole\\CS-1A\\assos\\automatant\\hackathon-GenIA-31-01-2025\\data\\Bail_Harmonie.pdf"
test_pdf_path_2 = "C:\\Users\\mateo\\Documents\\ecole\\CS-1A\\assos\\automatant\\hackathon-GenIA-31-01-2025\\data\\Bail_BUREAURO.pdf"
excel_path = "Tableau_audit_immo_modele.xlsx"
output_file_path = 'generated.xlsx'

def populate_line(wb, sheet_name, row, dict, correspondance):
    j = 1
    for elt in correspondance:
        j += 1
        wb[sheet_name].cell(row=row, column=j).value = dict[correspondance[j-2]]

def populate_excel(excel_filename, out_path, pdfs_path):
    wb = load_workbook(filename=excel_filename)
    i = 3
    for pdf_path in pdfs_path:
        i += 1
        print("processing", pdf_path)
        with open(pdf_path, "rb") as f:
            dict = extract_from_pdf(f, excel_extraction_page_charge + excel_extraction_page_duree + excel_extraction_page_divers)
        dict["EMPTY"] = ""
        populate_line(wb, "Durée - Loyers", i, dict, correspondance_duree)
        populate_line(wb, "Refacturation des charges", i, dict, correspondance_charges)
        populate_line(wb, "Divers", i, dict, correspondance_divers)
    wb.save(out_path)

def populate_excel_fast(excel_filename, out_path, pdfs_path, max_simult=10):
    wb = load_workbook(filename=excel_filename)
    i = 3
    for pdf_path in pdfs_path:
        i += 1
        print("processing", pdf_path)
        with open(pdf_path, "rb") as f:
            dict = extract_from_pdf_fast(f, excel_extraction_page_charge + excel_extraction_page_duree + excel_extraction_page_divers, max_simult)
        dict["EMPTY"] = ""
        populate_line(wb, "Durée - Loyers", i, dict, correspondance_duree)
        populate_line(wb, "Refacturation des charges", i, dict, correspondance_charges)
        populate_line(wb, "Divers", i, dict, correspondance_divers)
    wb.save(out_path)

if __name__=="__main__":
    populate_excel_fast(excel_path, output_file_path,[test_pdf_path, test_pdf_path_2], 50)
        